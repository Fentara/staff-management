import openpyxl
from openpyxl.styles import Alignment, PatternFill, Border, Side, Font

def write_table_data(sheet, row_index, data, start_column=1, decimal_places=2, cell_color=None, fill_type='solid', border=True, font=None, horizontal_alignment='left', vertical_alignment='top'):
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    number_formats = {int: '0', float: f'0.{"0" * decimal_places}', str: '@'}

    for col_index, value in enumerate(data, start=start_column):
        cell = sheet.cell(row=row_index, column=col_index, value=value)
        if cell_color:
            cell.fill = PatternFill(start_color=cell_color, end_color=cell_color, fill_type=fill_type)
        cell.number_format = number_formats.get(type(value), '@')
        if border:
            cell.border = thin_border
        if font:
            font_styles = {
                "bold": "bold" in font,
                "italic": "italic" in font,
                "underline": "single" if "underline" in font else "none",
                "size": int(next((s.split("=")[1] for s in font.split() if s.startswith("size=")), 11)),
                "name": next((s.split("=")[1] for s in font.split() if s.startswith("name=")), "Calibri")}
            cell.font = Font(**font_styles)
        cell.alignment = Alignment(horizontal=horizontal_alignment, vertical=vertical_alignment)

def auto_adjust_column_width(sheet):
    for column_cells in sheet.columns:
        max_length = 0
        if isinstance(column_cells[0], openpyxl.cell.cell.Cell):
            column_letter = column_cells[0].column_letter
        else:
            continue
        header_cell = sheet.cell(row=1, column=column_cells[0].column)
        try:
            header_length = len(str(header_cell.value))
        except:
            header_length = 0
        if header_length > max_length:
            max_length = header_length
        for cell in column_cells:
            try:
                if cell.value:
                    length = len(str(cell.value))
                    if length > max_length:
                        max_length = length
            except:
                pass
        adjusted_width = max((max_length, 1))
        sheet.column_dimensions[column_letter].width = adjusted_width

def title_merge_format(sheet, start_row, start_col, end_row, end_col, content=None, border=False):
    sheet.cell(row=start_row, column=start_col).value = content
    sheet.merge_cells(start_row=start_row, start_column=start_col, end_row=end_row, end_column=end_col)
    cell = sheet.cell(row=start_row, column=start_col)
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.font = openpyxl.styles.Font(bold=True, size=14)
    if border:
        cell.border = openpyxl.styles.Border(left=openpyxl.styles.Side(style='thin'), right=openpyxl.styles.Side(style='thin'), top=openpyxl.styles.Side(style='thin'), bottom=openpyxl.styles.Side(style='thin'))