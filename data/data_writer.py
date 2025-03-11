"""
Module for writing data analysis results to Excel files.

This module creates an Excel workbook with analysis results from various data sources.
"""

import openpyxl
from utils.file_paths import get_output_path
from data.data_analysis import (
    program_fte, program_area, program_mismatches, program_matches,
    program_totals_by_psych, school_totals_by_psych, total_psych_fte,
    psychs_for_program_types, psych_portfolios, school_distances
)


def write_data(staff_list, program_type_list, school_list, program_list):
    """Create an Excel workbook and populate it with analysis data.
    
    Args:
        staff_list (list): List of Staff objects
        program_type_list (list): List of ProgramType objects
        school_list (list): List of School objects
        program_list (list): List of Program objects
        
    Returns:
        None: The data is written to an Excel file at the path from get_output_path()
    """
    output_path = get_output_path()
    output = openpyxl.Workbook()

    output.create_sheet("Psych Analysis")
    output.create_sheet("Program Analysis")
    output.create_sheet("Psychs for Each Program")
    output.create_sheet("Psych Portfolios")
    output.create_sheet("School Distances")
    output.remove(output["Sheet"])

    # Call your analysis functions here
    # Example: program_fte(output["Program Analysis"], staff_list, program_type_list, school_list, program_list)

 # Call analysis functions with required parameters
    program_fte(output["Program Analysis"], staff_list, program_type_list, school_list, program_list, output_path)
    program_area(output["Program Analysis"], staff_list, program_type_list, school_list, program_list, output_path)
    program_mismatches(output["Program Analysis"], staff_list, program_type_list, school_list, program_list, output_path)
    program_matches(output["Program Analysis"], staff_list, program_type_list, school_list, program_list, output_path)
    program_totals_by_psych(output["Psych Analysis"], staff_list, program_type_list, school_list, program_list, output_path)
    school_totals_by_psych(output["Psych Analysis"], staff_list, program_type_list, school_list, program_list, output_path)
    total_psych_fte(output["Psych Analysis"], staff_list, program_type_list, school_list, program_list, output_path)
    psychs_for_program_types(output["Psychs for Each Program"], staff_list, program_type_list, school_list, program_list, output_path)
    psych_portfolios(output["Psych Portfolios"], staff_list, program_type_list, school_list, program_list, output_path)
    school_distances(output["School Distances"], staff_list, program_type_list, school_list, program_list, output_path)


    # Save the workbook
    output.save(output_path)
    print(f"Data written to {output_path}")