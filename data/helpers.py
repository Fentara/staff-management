"""
Helper functions for Excel operations using openpyxl.

This module provides utility functions for writing formatted data to Excel sheets,
calculating distances between schools, and formatting Excel tables.
"""

import math
import openpyxl
from openpyxl.styles import Alignment, PatternFill, Border, Side, Font

# Constants used across analysis functions
AREAS = ["NW", "NE", "CENTRAL", "SW", "SE"]
TABLE_COLORS = {
    "Light green": "C6EFCE",        # Light green
    "Dark green": "1AAD5A",  # Dark green
    "Light blue": "DBE0FF",          # Light blue
    "Dark blue": "848699",    # Dark blue
    "Light red": "FFC7CE",       # Light red
    "Dark red": "CD5C5C"  # Dark red
}

def haversine_distance(school1, school2):
    """Calculates the distance between two schools using the Haversine formula, to two decmial points."""
    try:
        sin_lats = math.sin(school1.get_latitude_radian()) * math.sin(school2.get_latitude_radian())
        cos_lats = math.cos(school1.get_latitude_radian()) * math.cos(school2.get_latitude_radian())
        cos_lons = math.cos(school1.get_longitude_radian() - school2.get_longitude_radian())
        SPHERE = 180 / math.pi * 60 * 1.852
        return round(math.acos(sin_lats + cos_lats * cos_lons) * SPHERE, 2)
    except (ValueError, TypeError) as e:
        print(f"Error: Unable to calculate distance: {e}")
        return None  

def write_table_data(sheet, row_index, data, start_column=1, decimal_places=2, cell_color=None, fill_type='solid', border=True, font=None, horizontal_alignment='left', vertical_alignment='top'):
    """Write data to an Excel sheet with formatting.

    This function writes data to cells in an Excel worksheet and applies formatting 
    options such as colors, borders, number formats, and alignment.

    Args:
        sheet (Worksheet): The Excel worksheet to write to.
        row_index (int): The row number to write data to (1-based).
        data (list): List of values to write to the row.
        start_column (int, optional): Column to start writing from. Defaults to 1.
        decimal_places (int, optional): Number of decimal places for float values. Defaults to 2.
        cell_color (str, optional): Hex color code for cell background. Defaults to None.
        fill_type (str, optional): Style of fill ('solid', 'gradient', etc.). Defaults to 'solid'.
        border (bool, optional): Whether to add borders to cells. Defaults to True.
        font (str, optional): Font formatting string (e.g., "bold size=12"). Defaults to None.
        horizontal_alignment (str, optional): Text horizontal alignment. Defaults to 'left'.
        vertical_alignment (str, optional): Text vertical alignment. Defaults to 'top'.
    
    Returns:
        None: This function doesn't return a value.
    """
    
    # Define border style
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'), 
        top=Side(style='thin'), bottom=Side(style='thin'))

    # Define number formats
    number_formats = {
        int: '0',
        float: f'0.{"0" * decimal_places}',
        str: '@'}

    # Cell by cell formatting (color, border, font, etc)
    for col_index, value in enumerate(data, start=start_column):
        cell = sheet.cell(row=row_index, column=col_index, value=value)

        # Apply fill color if provided
        if cell_color:
            cell.fill = PatternFill(start_color=cell_color, end_color=cell_color, fill_type=fill_type)

        # Apply number format based on value type
        cell.number_format = number_formats.get(type(value), '@')

        # Apply border if enabled
        if border:
            cell.border = thin_border

        # Apply font styles if specified
        if font:
            font_styles = {
                "bold": "bold" in font,
                "italic": "italic" in font,
                "underline": "single" if "underline" in font else "none",
                "size": int(next((s.split("=")[1] for s in font.split() if s.startswith("size=")), 11)),
                "name": next((s.split("=")[1] for s in font.split() if s.startswith("name=")), "Calibri")}
            cell.font = Font(**font_styles)
        
        # Apply alignment
        cell.alignment = Alignment(horizontal=horizontal_alignment, vertical=vertical_alignment)

def find_first_empty_column(sheet):
    """Find the first empty column in the given sheet."""
    for col in range(1, sheet.max_column + 2):
        if all(sheet.cell(row=row, column=col).value is None for row in range(1, sheet.max_row + 1)):
            return col
    return sheet.max_column + 1

def find_last_filled_column(sheet):
    """Find the final column with data in the given sheet."""
    for col in range(1, sheet.max_column + 2):
        return sheet.max_column + 1  # If no empty column is found, return the next column

def auto_adjust_column_width(sheet):
    for column_cells in sheet.columns:
        max_length = 0

        # Check if the first cell is a merged cell
        if isinstance(column_cells[0], openpyxl.cell.cell.Cell):
            column_letter = column_cells[0].column_letter
        else:
            continue  # Skip merged cells

        # Get the header cell
        header_cell = sheet.cell(row=1, column=column_cells[0].column)

        # Calculate the length of the header
        try:
            header_length = len(str(header_cell.value))
        except:
            header_length = 0

        # Update max_length if the header is longer
        if header_length > max_length:
            max_length = header_length

        for cell in column_cells:
            try:
                if cell.value:
                    length = len(str(cell.value))
                    if length > max_length:
                        max_length = length
            except:
                pass

        adjusted_width = max((max_length, 1))  # Minimum width of 1
        
        sheet.column_dimensions[column_letter].width = adjusted_width

def title_merge_format(sheet, start_row, start_col, end_row, end_col, content=None, border = False):
    """Merge and format cells in the given range."""
    # Safety check - Excel uses 1-based indexing
    start_row = max(1, start_row)
    start_col = max(1, start_col)
    end_row = max(1, end_row)
    end_col = max(1, end_col)
                  
    sheet.cell(row=start_row, column=start_col).value = content
    sheet.merge_cells(start_row=start_row, start_column=start_col, end_row=end_row, end_column=end_col)
    cell = sheet.cell(row=start_row, column=start_col)
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.font = openpyxl.styles.Font(bold=True, size = 14)
    if border:
        cell.border = openpyxl.styles.Border(left=openpyxl.styles.Side(style='thin'), right=openpyxl.styles.Side(style='thin'), top=openpyxl.styles.Side(style='thin'), bottom=openpyxl.styles.Side(style='thin'))

def get_psychologists(staff_list):
    """Return a list of staff members who are psychologists."""
    return [p for p in staff_list if p.get_job().casefold() == "psychologist"]

def create_area_dict():
    """Create a dictionary with area keys initialized to zero."""
    return {area: 0 for area in AREAS}

def write_table_header(sheet, row, headers, start_col=1, color_key="Light green"):
    """Write a table header with consistent formatting.
    
    Args:
        sheet: The worksheet to write to
        row: The row number to write the header to
        headers: List of header values
        start_col: Starting column (default: 1)
        color_key: Key to use for TABLE_COLORS (default: "program")
    """
    write_table_data(
        sheet, row, headers, 
        start_column=start_col,
        cell_color=TABLE_COLORS.get(color_key, "C6EFCE"), 
        font="bold", 
        horizontal_alignment="center", 
        vertical_alignment="center"
    )
    
def write_total_row(sheet, row, data, start_col=1, color_key="dark green"):
    """Write a total row with consistent formatting.
    
    Args:
        sheet: The worksheet to write to
        row: The row number to write the total row to
        data: List of values for the total row
        start_col: Starting column (default: 1)
        color_key: Key to use for TABLE_COLORS (default: "program_total")
    """
    write_table_data(
        sheet, row, data, 
        start_column=start_col, 
        cell_color=TABLE_COLORS.get(color_key, "1AAD5A"), 
        font="bold"
    )

def format_table(sheet, start_row, start_col, end_col, title):
    """Apply common table formatting including column width adjustment and title.
    
    Args:
        sheet: The worksheet to format
        start_row: Row where the title should be placed
        start_col: Starting column of the title merge
        end_col: Ending column of the title merge
        title: Text for the title
    """
    # Safety check - Excel uses 1-based indexing
    start_row = max(1, start_row)
    start_col = max(1, start_col)
    end_col = max(1, end_col)

    auto_adjust_column_width(sheet)
    title_merge_format(sheet, start_row, start_col, start_row, end_col, content=title)