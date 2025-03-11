import openpyxl
from utils.file_paths import get_paths, get_output_path
from models.staff import Staff
from models.program_type import ProgramType
from models.school import School
from models.program import Program

SHEETS = {
    "STAFF": "staff",
    "PROGRAM_TYPES": "program_types",
    "SCHOOLS": "schools",
    "PROGRAMS": "programs"
}

COLUMNS = {
    "STAFF": {
        "NAME": 1,
        "JOB": 2,
        "FTE": 3,
        "TEAM": 4,
        "SPED_PROGRAMS": 5,
        "BEH_PROGRAMS": 6,
        "AREA": 7
    },
    "PROGRAM_TYPES": {
        "NAME": 1,
        "TEAM": 2,
        "ADAPTIVE_FUNC": 3,
        "COGNITIVE_FUNC": 4,
        "SOC_EMO_BEH_FUNC": 5,
        "PHYS_MED_NEED": 6,
        "WEIGHT": 7
    },
    "SCHOOLS": {
        "NAME": 1,
        "AREA": 2,
        "SCHOOL_PSYCH": 3,
        "ADDRESS": 4,
        "LATITUDE": 7,
        "LONGITUDE": 8,
        "GRADES": 9
    },
    "PROGRAMS": {
        "SCHOOL": 1,
        "PROGRAM_TYPE": 3,
        "PSYCH": 4
    }
}


def load_data():
    try:
        primary_path, alternate_path, path = get_paths()
        wb_obj = openpyxl.load_workbook(path)
        
        staff_list = create_staff(wb_obj)
        program_type_list = create_program_types(wb_obj)
        school_list = create_schools(wb_obj)
        program_list = create_programs(wb_obj)
        
        return staff_list, program_type_list, school_list, program_list
    except FileNotFoundError:
        print(f"Error: Excel file not found at {path}")
        return [], [], [], []
    except KeyError as e:
        print(f"Error: Sheet {e} not found in workbook")
        return [], [], [], []

def create_staff(wb_obj):
    """Create Staff objects from the 'staff' sheet in the workbook.
    
    Args:
        wb_obj (openpyxl.workbook.Workbook): The loaded Excel workbook
        
    Returns:
        list: A list of Staff objects
    """

    staff_sheet = wb_obj[SHEETS["STAFF"]]
    staff_list = []
    for i in range(2, staff_sheet.max_row + 1):
        name = staff_sheet.cell(row=i, column=COLUMNS["STAFF"]["NAME"]).value
        job = staff_sheet.cell(row=i, column=COLUMNS["STAFF"]["JOB"]).value
        fte = staff_sheet.cell(row=i, column=COLUMNS["STAFF"]["FTE"]).value
        team = staff_sheet.cell(row=i, column=COLUMNS["STAFF"]["TEAM"]).value
        sped_programs = staff_sheet.cell(row=i, column=COLUMNS["STAFF"]["SPED_PROGRAMS"]).value
        beh_programs = staff_sheet.cell(row=i, column=COLUMNS["STAFF"]["BEH_PROGRAMS"]).value

        sped_programs = sped_programs.split(', ') if sped_programs else []
        beh_programs = beh_programs.split(', ') if beh_programs else []

        staff_member = Staff(name, job, fte, team, sped_programs, beh_programs)
        staff_list.append(staff_member)
    return staff_list

def create_program_types(wb_obj):
    """Create ProgramType objects from the 'program_types' sheet in the workbook.
    
    Args:
        wb_obj (openpyxl.workbook.Workbook): The loaded Excel workbook
        
    Returns:
        list: A list of ProgramType objects
    """

    program_type_sheet = wb_obj['program_types']
    program_type_list = []
    for i in range(2, program_type_sheet.max_row + 1):
        name = program_type_sheet.cell(row=i, column=COLUMNS["PROGRAM_TYPES"]["NAME"]).value
        team = program_type_sheet.cell(row=i, column=COLUMNS["PROGRAM_TYPES"]["TEAM"]).value
        adaptive_func = program_type_sheet.cell(row=i, column=COLUMNS["PROGRAM_TYPES"]["ADAPTIVE_FUNC"]).value
        cognitive_func = program_type_sheet.cell(row=i, column=COLUMNS["PROGRAM_TYPES"]["COGNITIVE_FUNC"]).value
        soc_emo_beh_func = program_type_sheet.cell(row=i, column=COLUMNS["PROGRAM_TYPES"]["SOC_EMO_BEH_FUNC"]).value
        phys_med_need = program_type_sheet.cell(row=i, column=COLUMNS["PROGRAM_TYPES"]["PHYS_MED_NEED"]).value
        weight = program_type_sheet.cell(row=i, column=COLUMNS["PROGRAM_TYPES"]["WEIGHT"]).value

        program = ProgramType(name, team, adaptive_func, cognitive_func, soc_emo_beh_func, phys_med_need, weight)
        program_type_list.append(program)
    return program_type_list

def create_schools(wb_obj):
    """Create School objects from the 'schools' sheet in the workbook.
    
    Args:
        wb_obj (openpyxl.workbook.Workbook): The loaded Excel workbook
        
    Returns:
        list: A list of School objects
    """
    school_sheet = wb_obj['schools']
    program_sheet = wb_obj['programs']
    school_list = []
    for i in range(2, school_sheet.max_row + 1):
        name = school_sheet.cell(row=i, column=COLUMNS["SCHOOLS"]["NAME"]).value
        area = school_sheet.cell(row=i, column=COLUMNS["SCHOOLS"]["AREA"]).value
        school_psych = school_sheet.cell(row=i, column=COLUMNS["SCHOOLS"]["SCHOOL_PSYCH"]).value
        address = school_sheet.cell(row=i, column=COLUMNS["SCHOOLS"]["ADDRESS"]).value
        latitude_radian = school_sheet.cell(row=i, column=COLUMNS["SCHOOLS"]["LATITUDE"]).value
        longitude_radian = school_sheet.cell(row=i, column=COLUMNS["SCHOOLS"]["LONGITUDE"]).value

        programs = []
        for j in range(2, program_sheet.max_row + 1):
            program_school_name = program_sheet.cell(row=j, column=1).value
            if program_school_name == name:
                program_name = program_sheet.cell(row=j, column=3).value
                programs.append(program_name)

        school = School(name, area, school_psych, address, latitude_radian, longitude_radian, programs)
        school_list.append(school)
    return school_list

def create_programs(wb_obj):
    """Create Program objects from the 'programs' sheet in the workbook.
    
    Args:
        wb_obj (openpyxl.workbook.Workbook): The loaded Excel workbook
        
    Returns:
        list: A list of Program objects
    """
    program_sheet = wb_obj['programs']
    program_list = []
    for i in range(2, program_sheet.max_row + 1):
        school = program_sheet.cell(row=i, column=COLUMNS["PROGRAMS"]["SCHOOL"]).value
        program_type = program_sheet.cell(row=i, column=COLUMNS["PROGRAMS"]["PROGRAM_TYPE"]).value
        psych = program_sheet.cell(row=i, column=COLUMNS["PROGRAMS"]["PSYCH"]).value

        program = Program(school, program_type, psych)
        program_list.append(program)
    return program_list