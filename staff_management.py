import openpyxl, math, statistics, os
import openpyxl.styles
import openpyxl.workbook
from openpyxl.styles import Alignment, PatternFill, Border, Side, Font
from collections import defaultdict

primary_path = "C:/Users/david.williamson/OneDrive - Calgary Catholic School District/Dave/Python/CCSD Management/ccsd data.xlsx"
alternate_path = "C:/Users/Dave/OneDrive - Calgary Catholic School District/Dave/Python/CCSD Management/ccsd data.xlsx"

path = primary_path if os.path.exists(primary_path) else alternate_path
output_path = os.path.join(os.path.dirname(path), "CCSD Output.xlsx")

wb_obj = openpyxl.load_workbook(path)

## Prep Functions ##
# Create variables to store each sheet, and the number of rows and columns in each sheet
staff_sheet = wb_obj['staff']
staff_rows = staff_sheet.max_row
staff_columns = staff_sheet.max_column

program_type_sheet = wb_obj['program_types']
program_type_rows = program_type_sheet.max_row
program_type_columns = program_type_sheet.max_column

school_sheet = wb_obj['schools']
school_rows = school_sheet.max_row
school_columns = school_sheet.max_column

program_sheet = wb_obj['programs']
program_rows = program_sheet.max_row
program_columns = program_sheet.max_column

# Class Definition Functions
class Staff:
    """A class to represent a staff member."""
    def __init__(self, name, job, fte, team, sped_programs = [], beh_programs = []):
        self._name = name
        self._job = job
        self._fte = fte
        self._team = team
        self._sped_programs = sped_programs
        self._beh_programs = beh_programs
    
    def __str__(self):
        return self._name + " is a " + self._job + " on the " + self._team + " team. They have an FTE of " + str(self._fte) + ", and their assigned programs are " + str(self._sped_programs) + " and " + str(self._beh_programs) + "."
    
    def get_name(self):
        """Return the name of the staff member."""
        return self._name
    
    def get_job(self):
        """Return the job of the staff member."""
        return self._job
    
    def get_fte(self):
        """Return the FTE of the staff member."""
        return self._fte

    def get_team(self):
        """Return the team of the staff member."""
        return self._team
    
    def get_programs(self):
        """Return the programs the staff member is assigned to."""
        return self._sped_programs, self._beh_programs

    def set_program(self, program):
        """Add a program to a staff member's list of programs."""
        if program.team == "SPED":
            self._sped_programs.append(program)
        elif program.team == "Behaviour":
            self._beh_programs.append(program)

class ProgramType:
    def __init__(self, name, team, adaptive_func, cognitive_func, soc_emo_beh_func, phys_med_need, weight):
        self._team = team
        self._name = name
        self._adapt = adaptive_func
        self._cog = cognitive_func
        self._seb = soc_emo_beh_func
        self._phys_med = phys_med_need
        self._weight = weight
    
    def __str__(self):
        return "The " + str(self._name) + " program is managed by the " + str(self._team) + " team. It supports students with " + str(self._adapt) + " adaptive functioning deficits, " + str(self._cog) + " cognitive functioning deficits, and " + str(self._seb) + " social emotional behavior functioning deficts. It has an FTE weight of " + str(self._weight) + "."
   
    def get_name(self):
        """Return the name of the program."""
        return self._name
    
    def get_team(self):
        """Return the team of the program."""
        return self._team
    
    def get_adaptive_func(self):
        """Return the adaptive functioning level that the program supports."""
        return self._adapt

    def get_cog(self):
        """Return the cognitive functioning level that the program supports."""
        return self._cog
    
    def get_seb(self):
        """Return the social emotional behavioural functioning level that the program supports."""
        return self._seb

    def get_phys_med(self):
        """Return the physical medical needs that the program supports."""
        return self._phys_med

    def get_weight(self):
        """Return the FTE weight of the program."""
        return self._weight

class School:
    def __init__(self, name, area, school_psych, address, latitude_radian, longitude_radian, programs):
        self._name = name
        self._area = area
        self._school_psych = school_psych
        self._address = address
        self._latitude_radian = latitude_radian
        self._longitude_radian = longitude_radian     
        self._programs = programs

    def __str__(self):
        return self._name + " is located in " + self._area + ". The school psychologist is " + self._school_psych + ". The address is " + self._address + ". The latitude is " + str(self._latitude_radian) + " and the longitude is " + str(self._longitude_radian) + ". The programs offered are " + str(self._programs) + "."

    def get_name(self):
        """Return the name of the school."""
        return self._name
    
    def get_area(self):
        """Return the city area of the school."""
        return self._area
    
    def get_school_psych(self):
        """Return the school psychologist who is assigned to that school."""
        return self._school_psych
    
    def get_address(self):
        """Return the address of the school."""
        return self._address
    
    def get_latitude_radian(self):
        """Return the latitude of the school in radians."""
        return self._latitude_radian
    
    def get_longitude_radian(self):
        """Return the longitude of the school in radians."""
        return self._longitude_radian
    
    def get_programs(self):
        """Return the Diverse Learning programs offered by the school."""
        return self._programs

class Program:
    def __init__(self, school, program, psych):
        self._school = school
        self._program = program
        self._psych = psych

    def __str__(self):
        return "The " + self._school + " offers the " + self._program + " program. It is supported by " + self._psych + "."

    def get_school(self):
        """Return the school that offers the program."""
        return self._school
    
    def get_program(self):
        """Return the type of program that is offered by the school."""
        return self._program
    
    def get_psych(self):
        """Return the psychologist who supports the program."""
        return self._psych

# Class Instantiation Functions    
def create_staff():
    """Create a list of Staff objects with information populated from the excel sheet."""
    staff_list = []
    for i in range(2, staff_rows + 1):
        # Extract the staff data from the excel sheet
        name = staff_sheet.cell(row=i, column=1).value
        job = staff_sheet.cell(row=i, column=2).value
        fte = staff_sheet.cell(row=i, column=3).value
        team = staff_sheet.cell(row=i, column=4).value
        sped_programs = staff_sheet.cell(row=i, column=5).value
        beh_programs = staff_sheet.cell(row=i, column=6).value

        # Split the program strings by comma and convert to lists
        sped_programs = sped_programs.split(', ') if sped_programs else []
        beh_programs = beh_programs.split(', ') if beh_programs else []
        
        staff_member = Staff(name, job, fte, team, sped_programs, beh_programs)
        staff_list.append(staff_member)
    return staff_list

def create_program_types():
    """Create a list of ProgramType objects from the excel sheet."""
    program_type_list = []
    for i in range(2, program_type_rows + 1):
        # Extract the program data from the excel sheet
        name = program_type_sheet.cell(row=i, column=1).value
        team = program_type_sheet.cell(row=i, column=2).value
        adaptive_func = program_type_sheet.cell(row=i, column=3).value
        cognitive_func = program_type_sheet.cell(row=i, column=4).value
        soc_emo_beh_func = program_type_sheet.cell(row=i, column=5).value
        phys_med_need = program_type_sheet.cell(row=i, column=6).value
        weight = program_type_sheet.cell(row=i, column=7).value

        program = ProgramType(name, team, adaptive_func, cognitive_func, soc_emo_beh_func, phys_med_need, weight)
        program_type_list.append(program)
    return program_type_list

def create_schools():
    """Create a list of School objects from the excel sheet."""
    school_list = []
    for i in range(2, school_rows + 1):
        # Extract the school data from the excel sheet
        name = school_sheet.cell(row=i, column=1).value
        area = school_sheet.cell(row=i, column=2).value
        school_psych = school_sheet.cell(row=i, column=3).value
        address = school_sheet.cell(row=i, column=4).value
        latitude_radian = school_sheet.cell(row=i, column=7).value
        longitude_radian = school_sheet.cell(row=i, column=8).value
        
        # Create a list of programs offered by the school
        programs = []
        for j in range(2, program_rows + 1):
            program_school_name = program_sheet.cell(row=j, column=1).value
            if program_school_name == name:
                program_name = program_sheet.cell(row=j, column=3).value
                programs.append(program_name)
        
        school = School(name, area, school_psych, address, latitude_radian, longitude_radian, programs)
        school_list.append(school)
    return school_list

def create_programs():
    """Create a list of Program objects from the excel sheet."""
    program_list = []
    for i in range(2, program_rows + 1):
        # Extract the program data from the excel sheet
        school = program_sheet.cell(row=i, column=1).value
        program_type = program_sheet.cell(row=i, column=3).value
        psych = program_sheet.cell(row=i, column=4).value

        program = Program(school, program_type, psych)
        program_list.append(program)
    return program_list

# Helper Functions
def haversine_distance(school1, school2):
    """Calculates the distance between two schools using the Haversine formula, to two decmial points."""
    sin_lats = math.sin(school1.get_latitude_radian()) * math.sin(school2.get_latitude_radian())
    cos_lats = math.cos(school1.get_latitude_radian()) * math.cos(school2.get_latitude_radian())
    cos_lons = math.cos(school1.get_longitude_radian() - school2.get_longitude_radian())
    SPHERE = 180 / math.pi * 60 * 1.852
    return round(math.acos(sin_lats + cos_lats * cos_lons) * SPHERE, 2)

def write_table_data(sheet, row_index, data, start_column=1, decimal_places=2, cell_color=None, fill_type='solid', border=True, font=None, horizontal_alignment='left', vertical_alignment='top'):
    """Helper function to write data to the sheet starting from a specified column."""
    
    # Define border style
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'), 
        top=Side(style='thin'), bottom=Side(style='thin'))

    # Define number formats
    number_formats = {
        int: '0',
        float: f'0.{"0" * decimal_places}',
        str: '@'}

    # Cell by cell formatting (color, border, font, etc)
    for col_index, value in enumerate(data, start=start_column):
        cell = sheet.cell(row=row_index, column=col_index, value=value)

        # Apply fill color if provided
        if cell_color:
            cell.fill = PatternFill(start_color=cell_color, end_color=cell_color, fill_type=fill_type)

        # Apply number format based on value type
        cell.number_format = number_formats.get(type(value), '@')

        # Apply border if enabled
        if border:
            cell.border = thin_border

        # Apply font styles if specified
        if font:
            font_styles = {
                "bold": "bold" in font,
                "italic": "italic" in font,
                "underline": "single" if "underline" in font else "none",
                "size": int(next((s.split("=")[1] for s in font.split() if s.startswith("size=")), 11)),
                "name": next((s.split("=")[1] for s in font.split() if s.startswith("name=")), "Calibri")}
            cell.font = Font(**font_styles)
        
        # Apply alignment
        cell.alignment = Alignment(horizontal=horizontal_alignment, vertical=vertical_alignment)

def find_first_empty_column(sheet):
    """Find the first empty column in the given sheet."""
    for col in range(1, sheet.max_column + 2):
        if all(sheet.cell(row=row, column=col).value is None for row in range(1, sheet.max_row + 1)):
            return col
    return sheet.max_column + 1

def find_last_filled_column(sheet):
    """Find the final column with data in the given sheet."""
    for col in range(1, sheet.max_column + 2):
        return sheet.max_column + 1  # If no empty column is found, return the next column

def auto_adjust_column_width(sheet):
    for column_cells in sheet.columns:
        max_length = 0

        # Check if the first cell is a merged cell
        if isinstance(column_cells[0], openpyxl.cell.cell.Cell):
            column_letter = column_cells[0].column_letter
        else:
            continue  # Skip merged cells

        # Get the header cell
        header_cell = sheet.cell(row=1, column=column_cells[0].column)

        # Calculate the length of the header
        try:
            header_length = len(str(header_cell.value))
        except:
            header_length = 0

        # Update max_length if the header is longer
        if header_length > max_length:
            max_length = header_length

        for cell in column_cells:
            try:
                if cell.value:
                    length = len(str(cell.value))
                    if length > max_length:
                        max_length = length
            except:
                pass

        adjusted_width = max((max_length, 1))  # Minimum width of 10
        
        sheet.column_dimensions[column_letter].width = adjusted_width

def title_merge_format(sheet, start_row, start_col, end_row, end_col, content=None, border = False):
    """Merge and format cells in the given range."""
    sheet.cell(row=start_row, column=start_col).value = content
    sheet.merge_cells(start_row=start_row, start_column=start_col, end_row=end_row, end_column=end_col)
    cell = sheet.cell(row=start_row, column=start_col)
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.font = openpyxl.styles.Font(bold=True, size = 14)
    if border:
        cell.border = openpyxl.styles.Border(left=openpyxl.styles.Side(style='thin'), right=openpyxl.styles.Side(style='thin'), top=openpyxl.styles.Side(style='thin'), bottom=openpyxl.styles.Side(style='thin'))

## Data Analysis Functions ##
# Functions to Summarize Programs #
def program_fte(sheet): 
    """Write the number and weighted FTE required for each program type to the program analysis sheet."""
    total_schools = len(school_list)

    # Count occurrences of each program type
    program_counts = {program.get_program(): 0 for program in program_list}
    for program in program_list:
        program_counts[program.get_program()] += 1

    # Write headers
    headers = ["Program Type", "Quantity", "FTE Weight", "Total FTE Needed"]
    write_table_data(sheet, 2, headers, cell_color='C6EFCE', font = "bold", horizontal_alignment="center", vertical_alignment="center")

    # Initialize totals
    sum_total_fte = sum_total_quantity = 0
    row_index = 3

    # Write program data
    for program in program_type_list:
        name = program.get_name()
        quantity = (
            total_schools if name == "Schools" 
            else 1 if name == "COPE" 
            else program_counts.get(name, 0)
        )
        
        fte_weight = program.get_weight()
        total_fte = quantity * fte_weight
        sum_total_fte += total_fte
        sum_total_quantity += quantity

        write_table_data(sheet, row_index, [name, quantity, fte_weight, total_fte], cell_color='C6EFCE')
        row_index += 1

    # Write total row
    write_table_data(sheet, row_index, ["Total", sum_total_quantity, "", sum_total_fte], cell_color='1AAD5A', font = "bold")
    
    # Format the table
    auto_adjust_column_width(sheet)
    title_merge_format(sheet, 1, 1, 1, len(headers), content = "Total District Program Summary")

    sheet.parent.save(output_path)

def program_area(sheet):    
    """Write the number of programs in each area to the program analysis sheet."""

    # Write headers
    headers = ["Program Type", "NW", "NE", "Central", "SW", "SE", "Total"]
    start_col = find_first_empty_column(sheet) + 1
    write_table_data(sheet, 2, headers, start_column=start_col, cell_color="C6EFCE", font="bold", horizontal_alignment="center", vertical_alignment="center")

    # Initialize area totals
    area_totals = {area: 0 for area in ["NW", "NE", "CENTRAL", "SW", "SE"]}
    program_type_counts = {p.get_name(): area_totals.copy() for p in program_type_list}

    # Count programs per area
    for program in program_list:
        school = next((s for s in school_list if s.get_name() == program.get_school()), None)
        if school:
            program_type_counts.setdefault(program.get_program(), area_totals.copy())[school.get_area()] += 1

    # Write program data to the sheet
    row_index = 3
    grand_total = 0
    for program_type, area_counts in program_type_counts.items():
        total_count = sum(area_counts.values())
        grand_total += total_count
        for area in area_totals:
            area_totals[area] += area_counts[area]
        write_table_data(sheet, row_index, [program_type, *area_counts.values(), total_count], start_column=start_col, cell_color="C6EFCE")
        row_index += 1

    # Count and write school totals
    schools_area_counts = {area: sum(1 for s in school_list if s.get_area() == area) for area in area_totals}
    total_schools_count = sum(schools_area_counts.values())
    grand_total += total_schools_count

    write_table_data(sheet, row_index, ["Schools", *schools_area_counts.values(), total_schools_count], start_column=start_col, cell_color="C6EFCE")
    row_index += 1

    # Write overall total row
    write_table_data(sheet, row_index, ["Total", *area_totals.values(), grand_total], start_column=start_col, cell_color="1AAD5A", font="bold")

    # Format the table
    auto_adjust_column_width(sheet)
    title_merge_format(sheet, 1, start_col, 1, start_col+len(headers)-1, content = "District Program Summary by Area")

    sheet.parent.save(output_path)

def program_mismatches(sheet):
    """Calculates and identifies how many and which programs are supported by psychologists who are not the school psych"""
    # Write headers
    headers = ["Program", "School", "Program Psychologist", "School Psychologist"]
    start_col = find_last_filled_column(sheet) + 1
    write_table_data(sheet, 2, headers, start_column=start_col, cell_color="FFC7CE", font="bold", horizontal_alignment="center", vertical_alignment="center")

    # Initialize totals
    row_index = 3
    mismatches = 0

    for program in program_list:
        school = next((s for s in school_list if s.get_name() == program.get_school()), None)
        if school:
            program_psych = next((s for s in staff_list if s.get_name() == program.get_psych()), None)
            if program_psych and program_psych.get_team() != "Behaviour":
                if program.get_psych() != school.get_school_psych():
                    write_table_data(sheet, row_index, [program.get_program(), program.get_school(), program.get_psych(), school.get_school_psych()], start_column=start_col, cell_color="FFC7CE")
                    row_index += 1
                    mismatches += 1

    # Write total row
    write_table_data(sheet, row_index, ["Total", "", "", mismatches], start_column=start_col, cell_color="CD5C5C", font="bold")

    # Format the table
    auto_adjust_column_width(sheet)
    title_merge_format(sheet, 1, start_col, 1, start_col+len(headers), content="Programs Psych/School Psych Mismatches")
    
    # Save the workbook
    sheet.parent.save(output_path)

def program_matches(sheet):
    """Calculates and identifies how many and which programs are supported by psychologists who are the school psych"""
    # Write headers
    headers = ["Program", "School", "Program Psychologist", "School Psychologist"]
    start_col = find_last_filled_column(sheet)
    write_table_data(sheet, 2, headers, start_column=start_col, cell_color="C6EFCE", font="bold", horizontal_alignment="center", vertical_alignment="center")

    # Initialize totals
    row_index = 3
    mismatches = 0

    for program in program_list:
        school = next((s for s in school_list if s.get_name() == program.get_school()), None)
        if school:
            program_psych = next((s for s in staff_list if s.get_name() == program.get_psych()), None)
            if program_psych and program_psych.get_team() != "Behaviour":
                if program.get_psych() == school.get_school_psych():
                    write_table_data(sheet, row_index, [program.get_program(), program.get_school(), program.get_psych(), school.get_school_psych()], start_column=start_col, cell_color="C6EFCE")
                    row_index += 1
                    mismatches += 1

    # Write total row
    write_table_data(sheet, row_index, ["Total", "", "", mismatches], start_column=start_col, cell_color="1AAD5A", font="bold")

    # Format the table
    auto_adjust_column_width(sheet)
    title_merge_format(sheet, 1, start_col, 1, start_col+len(headers), content="Programs Psych/School Psych Matches")

# Functions to Summarize Psychologists #
def program_totals_by_psych(sheet):
    """Writes the number of programs each psychologist supports in each area and their total program FTE to the psych analysis sheet."""

    # Write headers
    headers = ["Psychologist", "NW", "NE", "CENTRAL", "SW", "SE", "Total Programs", "Total Program FTE"]
    write_table_data(sheet, 2, headers, cell_color="DBE0FF", font="bold", horizontal_alignment="center", vertical_alignment="center")

    # Create lookup dictionaries for efficiency
    school_dict = {s.get_name(): s for s in school_list}
    program_dict = {pt.get_name(): pt for pt in program_type_list}

    row_index = 3
    psychologists = [p for p in staff_list if p.get_job().casefold() == "psychologist"]

    # Initialize totals
    total_area_counts = defaultdict(int)
    grand_total_programs = grand_total_fte = 0

    for psych in psychologists:
        psych_name = psych.get_name()
        area_counts = defaultdict(int)
        total_programs = total_fte = 0

        for program in program_list:
            if program.get_psych() == psych_name:
                school = school_dict.get(program.get_school())
                if school:
                    area_counts[school.get_area()] += 1
                    total_programs += 1
                    program_type = program_dict.get(program.get_program())
                    if program_type:
                        total_fte += program_type.get_weight()

        data = [
            psych_name,
            area_counts["NW"],
            area_counts["NE"],
            area_counts["CENTRAL"],
            area_counts["SW"],
            area_counts["SE"],
            total_programs,
            total_fte]
        write_table_data(sheet, row_index, data, cell_color="DBE0FF")
        row_index += 1

        # Update totals
        for area, count in area_counts.items():
            total_area_counts[area] += count
        grand_total_programs += total_programs
        grand_total_fte += total_fte

    # Write Total Row
    total_data = [
        "Total",
        total_area_counts["NW"],
        total_area_counts["NE"],
        total_area_counts["CENTRAL"],
        total_area_counts["SW"],
        total_area_counts["SE"],
        grand_total_programs,
        grand_total_fte]
    write_table_data(sheet, row_index, total_data, cell_color="848699", font="bold")
    
    # Format the table
    auto_adjust_column_width(sheet)
    title_merge_format(sheet, 1, 1, 1, len(headers), content = "Psych Program Support by Area")

    # Save the workbook
    sheet.parent.save(output_path)

def school_totals_by_psych(sheet):
    """Writes the number of schools each psychologist supports in each area to the psych analysis sheet."""

    # Determine the starting column
    start_col = find_first_empty_column(sheet) + 1

    # Write headers
    headers = ["Psychologist", "NW", "NE", "CENTRAL", "SW", "SE", "Total Schools", "Total School FTE"]
    write_table_data(sheet, 2, headers, start_column=start_col, cell_color="DBE0FF", font="bold", horizontal_alignment="center", vertical_alignment="center")

    row_index = 3
    psychologists = [p for p in staff_list if p.get_job().casefold() == "psychologist"]

    # Initialize totals
    total_area_counts = defaultdict(int)
    grand_total_schools = grand_total_weight = 0

    for psych in psychologists:
        psych_name = psych.get_name()
        area_counts = defaultdict(int)
        total_schools = total_weight = 0

        for program_type in program_type_list:
            if program_type.get_name() == "Schools":
                school_weight = program_type.get_weight()
                break

        for school in school_list:
            assigned_psych = school.get_school_psych()
            school_area = school.get_area()

            if assigned_psych == psych_name:
                area_counts[school_area] += 1
                total_schools += 1

        # Ensure at least one school is counted for the psychologist
        if total_schools > 0:
            total_weight = total_schools * school_weight
            data = [
                psych_name,
                area_counts["NW"],
                area_counts["NE"],
                area_counts["CENTRAL"],
                area_counts["SW"],
                area_counts["SE"],
                total_schools,
                total_weight]
            write_table_data(sheet, row_index, data, start_column=start_col, cell_color="DBE0FF")
            row_index += 1

            # Update totals
            for area, count in area_counts.items():
                total_area_counts[area] += count
            grand_total_schools += total_schools
            grand_total_weight += total_weight

    # Write Total Row
    total_data = [
        "Total",
        total_area_counts["NW"],
        total_area_counts["NE"],
        total_area_counts["CENTRAL"],
        total_area_counts["SW"],
        total_area_counts["SE"],
        grand_total_schools,
        grand_total_weight]
    write_table_data(sheet, row_index, total_data, start_column=start_col, cell_color="848699", font="bold")
    
    # Format the table
    auto_adjust_column_width(sheet)
    title_merge_format(sheet, 1, start_col, 1, start_col + len(headers) - 1, content="Psych School Support by Area")


    # Save the workbook
    sheet.parent.save(output_path)

def total_psych_fte(sheet):
    """Writes the total FTE of all psychologists to the psych analysis sheet."""

    # Determine the starting row and column
    row_index = 1
    start_col = find_last_filled_column(sheet) + 1

    # Write headers
    headers = ["Psychologist", "Total FTE"]
    write_table_data(sheet, 2, headers, start_column=start_col, cell_color="DBE0FF", font="bold", horizontal_alignment="center", vertical_alignment="center")

    # Create lookup dictionaries for efficiency
    program_dict = {pt.get_name(): pt for pt in program_type_list}

    # Calculate the total FTE of each psychologist
    row_index = 3
    psychologists = [p for p in staff_list if p.get_job().casefold() == "psychologist"]
    grand_total = 0

    for psych in psychologists:
        psych_name = psych.get_name()
        school_weight = next((pt.get_weight() for pt in program_type_list if pt.get_name() == "Schools"), 0)
        total_fte = 0

        for program in program_list:
            program_type = program_dict.get(program.get_program())
            if program.get_psych() == psych_name:
                if program_type:
                    total_fte += program_type.get_weight()
                    grand_total += program_type.get_weight()
        
        for school in school_list:
            if school.get_school_psych() == psych_name:
                total_fte += school_weight
                grand_total += school_weight

    # Write data to the sheet
        data = [psych_name, total_fte]
        write_table_data(sheet, row_index, data, start_column=start_col, cell_color="DBE0FF")
        row_index += 1
    
    # Write total row
    write_table_data(sheet, row_index, ["TOTAL", grand_total], start_column=start_col, cell_color="848699", font="bold")

    # Format the table
    auto_adjust_column_width(sheet)
    title_merge_format(sheet, 1, start_col, 1, start_col + len(headers)-1, "Total Assigned FTE per Psych")

    # Save the workbook
    sheet.parent.save(output_path)

# Functions for Individual Worksheets
def psychs_for_program_types(sheet):
    """Creates a sheet that lists all program types and writes all of the psychologists who support that program type."""
    
    # Write the program names across the first row
    program_names = [program.get_name() for program in program_type_list]
    write_table_data(sheet, 1, program_names, cell_color="DBE0FF", font="bold", horizontal_alignment="center", vertical_alignment="center")

    # Iterate over each program name and write the names of psychologists
    for col_index, program_name in enumerate(program_names, start=1):
        row_index = 2 
        for staff in staff_list:
            if staff.get_job().lower() == "psychologist":
                sped_programs, beh_programs = staff.get_programs()
                if program_name in sped_programs or program_name in beh_programs:
                    write_table_data(sheet, row_index, [staff.get_name()], start_column=col_index)
                    row_index += 1

    # Adjust column widths
    auto_adjust_column_width(sheet)

    # Save the workbook
    sheet.parent.save(output_path)

def psych_portfolios(sheet):
    """Creates a worksheet that lists all programs and all schools supported by each psychologist."""
    
    psych_names = [psych.get_name() for psych in staff_list if psych.get_job().lower() == "psychologist"]
    col_index = 1

    # Write "Schools" and "Programs" in the second row, alternating across columns
    col_index = 1
    while col_index <= len(psych_names) * 3:
        write_table_data(sheet, 2, ["Schools", "Programs"], start_column=col_index, cell_color="DBE0FF", font="bold", horizontal_alignment="center", vertical_alignment="center")
        col_index += 3  # Increment by 3 to leave a blank column between each psychologist

    # Write the list of schools in the "Schools" column for each psychologist
    for col_index, psych_name in enumerate(psych_names, start=1):
        row_index = 3  # Start writing schools from the third row
        for school in school_list:
            if school.get_school_psych() == psych_name:
                write_table_data(sheet, row_index, [school.get_name()], start_column=col_index * 3 - 2)
                row_index += 1

    # Write the list of programs in the "Programs" column for each psychologist
    for col_index, psych_name in enumerate(psych_names, start=1):
        row_index = 3  # Start writing programs from the third row
        for program in program_list:
            if program.get_psych() == psych_name:
                program_name = program.get_program()
                school_name = program.get_school()
                write_table_data(sheet, row_index, [f"{program_name} - {school_name}"], start_column=col_index * 3 - 1)
                row_index += 1

    # Adjust column widths
    auto_adjust_column_width(sheet)

    # Write the psychologist names across the first row
    col_index = 1
    for psych_name in psych_names:
        title_merge_format(sheet, 1, col_index, 1, col_index + 1, content=psych_name)
        col_index += 3  # Increment by 3 to leave a blank column between each psychologist

    # Save the workbook
    sheet.parent.save(output_path)
    print("Data written to CCSD Output.xlsx")

def school_distances(sheet):
    """Creates a worksheet calculating distances between schools in a psychologist's portfolio."""
    all_distances = []
    col_index = 1
    psych_distance_data = {}
    
    # Calculate the distance between each pair of schools in a psychologist's portfolio and write to the sheet (and median and max distance)
    for psych in [p for p in staff_list if p.get_job().lower() == "psychologist"]:
        psych_name = psych.get_name()
        psych_schools = [s for s in school_list if s.get_school_psych() == psych_name]
        row_index = 2
        distances = []

        for i, school1 in enumerate(psych_schools):
            for j in range(i + 1, len(psych_schools)):
                school2 = psych_schools[j]
                distance = haversine_distance(school1, school2)
                all_distances.append(distance)
                distances.append(distance)

                write_table_data(sheet, row_index, [f"{school1.get_name()} - {school2.get_name()}", distance], start_column=col_index)
                row_index += 1
        
        if not distances:
            continue

        # Calculate the median and maximum distance between schools in the portfolio and assign to psych_distance_data
        median_distance = statistics.median(distances)
        max_distance = max(distances)
        psych_distance_data[psych_name] = (median_distance, max_distance)

        write_table_data(sheet, row_index, ["Median Distance", median_distance], start_column=col_index)
        row_index += 1

        write_table_data(sheet, row_index, ["Max Distance", max_distance], start_column=col_index)

        # Write the psychologist's name in the first row
        title_merge_format(sheet, 1, col_index, 1, col_index + 1, content=psych_name)

        col_index += 3  # Increment by 3 to leave a blank column between each psychologist

    # Calculate the median and maximum distance between schools in all portfolios
    if all_distances:
        median_distance_all = statistics.median(all_distances)
        max_distance_all = max(all_distances)

    # Create Summary Table 
    summary_col_index = sheet.max_column + 2  # Ensures it appears after existing data
    write_table_data(sheet, 1, ["Psychologist", "Median Distance", "Max Distance"], start_column=summary_col_index, cell_color="DBE0FF", font="bold", horizontal_alignment="center", vertical_alignment="center")
    for row_index, (psych_name, (median_distance, max_distance)) in enumerate(psych_distance_data.items(), start=2):
        write_table_data(sheet, row_index, [psych_name, median_distance, max_distance], start_column=summary_col_index)
    write_table_data(sheet, row_index + 1, ["ALL Psychs", round(median_distance_all, 2), round(max_distance_all, 2)], start_column=summary_col_index, cell_color="848699", font="bold")

    # Adjust column widths
    auto_adjust_column_width(sheet)

    # Save the workbook
    sheet.parent.save(output_path)
    print("Data written to CCSD Output.xlsx")

## Main Program ##
## Create the staff, program, and school objects
staff_list = create_staff()
program_type_list = create_program_types()
school_list = create_schools()
program_list = create_programs()

# Create the workbook and sheets
output = openpyxl.Workbook()

# Establish sheets for analysis
output.create_sheet("Psych Analysis")
output.create_sheet("Program Analysis")
output.create_sheet("Psychs for Each Program")
output.create_sheet("Psych Portfolios")
output.create_sheet("School Distances")
output.remove(output["Sheet"])

# Populate the sheets with data
program_fte(output["Program Analysis"])
program_area(output["Program Analysis"])
program_mismatches(output["Program Analysis"])
program_matches(output["Program Analysis"])
program_totals_by_psych(output["Psych Analysis"])
school_totals_by_psych(output["Psych Analysis"])
total_psych_fte(output["Psych Analysis"])
psychs_for_program_types(output["Psychs for Each Program"])
psych_portfolios(output["Psych Portfolios"])
school_distances(output["School Distances"])
